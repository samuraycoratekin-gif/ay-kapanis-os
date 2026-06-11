# -*- coding: utf-8 -*-
"""
Akilli Mutabakat modulu — Ay Kapanis OS platformuna gomulu surum.

Mutabakat_AI/demo/app.py mantiginin tasinmis halidir. Sunucu/Handler/os.chdir
kaldirildi; tum yollar AKTIF KIRACIYA gore turetilir:
    veri/kiracilar/<kid>/mutabakat/
Platform Handler'i /mutabakat/* isteklerini dispatch_get/dispatch_post'a yonlendirir.
"""
import os, json, base64, threading, random, hashlib, hmac, smtplib, ssl, mimetypes
from datetime import datetime, timedelta
from email.message import EmailMessage
from email.utils import make_msgid
from urllib.parse import urlparse, parse_qs

from core import depo, kiraci, ayarlar, kasa
from . import motor as M, gib as G, durum as D

HERE = os.path.dirname(os.path.abspath(__file__))
TPL = os.path.join(HERE, "templates")

DONEM = "Mayıs 2026"
FIRMA = "ÖRNEK SANAYİ A.Ş."
GKEY = "__gonderim__"

OTP_DEPO = {}            # {"kid|musteri|cari": {"kod","bitis"}}  (bellekte, surec genelinde)
OTP_GECERLILIK = 10      # dakika

# Durum/kanit dosya yazimlari icin kilit (ThreadingHTTPServer altinda es zamanli
# istekler ayni JSON/JSONL'a yazabilir; hash zinciri ve durum bozulmasin).
_YAZ_KILIT = threading.RLock()


# --------------------------------------------------------------------------- #
# Kiraciya gore izole veri kokleri (eski sabit yol global'leri yerine fonksiyon)
# --------------------------------------------------------------------------- #
def _kok():
    k = os.path.join(depo.ROOT_VERI, "kiracilar", depo.aktif_kiraci(), "mutabakat")
    os.makedirs(k, exist_ok=True)
    return k


def _p(ad):
    return os.path.join(_kok(), ad)


def _uploads():
    u = _p("uploads")
    os.makedirs(u, exist_ok=True)
    return u


def DEMO_DURUM():   return _p("durum.json")
def MUKELLEFLER():  return _p("mukellefler.json")
def CARI_MAIL():    return _p("cari_mailler.json")
def MAIL_AYAR():    return _p("mail_ayar.json")
def KANIT_LOG():    return _p("kanit_log.jsonl")


# Repoyla birlikte gelen sentetik ornek ekstreler. Kiracinin kendi dosyasi
# yoksa bunlara duser: yeni kiraci (ve bulut) kurulumsuz calisir.
ORNEK_VERI = os.path.join(HERE, "ornek_veri")


def _ornek(anahtar):
    adlar = {"bizim": "bizim_ekstreler.xlsx", "karsi": "karsi_ekstreler.xlsx",
             "gib": "gib_efatura_kayitlari.xlsx"}
    kiraci_yol = os.path.join(_kok(), "ornek", adlar[anahtar])
    if os.path.exists(kiraci_yol):
        return kiraci_yol
    return os.path.join(ORNEK_VERI, adlar[anahtar])


def _yol_coz(yol, anahtar):
    if yol:
        p = yol if os.path.isabs(yol) else os.path.join(_kok(), yol)
        if os.path.exists(p):
            return p
        # Kiraci kopyasi yoksa ayni ada sahip paketlenmis ornege dus
        # (or. "ornek/muk015_bizim.xlsx" -> moduller/mutabakat/ornek_veri/muk015_bizim.xlsx)
        paket = os.path.join(ORNEK_VERI, os.path.basename(p))
        if os.path.exists(paket):
            return paket
        return p
    return _ornek(anahtar)


def _istek_host():
    """Bu istegi karsilayan Host basligi (app.py dispatch'te set edilir)."""
    return getattr(_CTX, "host", "") or ""


def PORTAL_URL():
    """Karsi tarafa giden linklerin koku. Oncelik: mail ayarindaki portal_url >
    istegin Host'u (bulutta https) > yerel varsayilan. Boylece ayar yapilmasa
    bile mail linki, ofisin kullandigi gercek adresi gosterir."""
    ayarli = (mail_ayar().get("portal_url") or "").strip()
    if ayarli:
        return ayarli.rstrip("/")
    host = _istek_host()
    if host:
        sema = "https" if os.environ.get("PORT") else "http"
        return f"{sema}://{host}/mutabakat"
    return "http://localhost:5050/mutabakat"


def _json_oku(yol, varsayilan):
    try:
        with open(yol, encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return varsayilan


def _json_yaz(yol, obj):
    """Atomik JSON yazimi (gecici dosya + replace)."""
    gecici = yol + ".tmp"
    with open(gecici, "w", encoding="utf-8") as f:
        json.dump(obj, f, ensure_ascii=False, indent=2)
    os.replace(gecici, yol)


def _kiraci_kayit():
    try:
        return kiraci.kiraci_getir(depo.aktif_kiraci()) or {}
    except Exception:
        return {}


def ofis_adi():
    """Mutabakat panosu basligi: ozellestirilmis ofis_adi varsa onu, yoksa kiraci unvanini kullanir.
    (Luca/ERP entegrasyonu geldiginde bu ad oradan da beslenebilir.)"""
    try:
        ad = (ayarlar.oku().get("ofis_adi") or "").strip()
    except Exception:
        ad = ""
    if ad and ad != "Ay Kapanış OS":
        return ad
    return (_kiraci_kayit().get("unvan") or ad or "Mutabakat").strip()


def kiraci_mail():
    """Kiracinin giris e-postasi — mutabakat formlarinda varsayilan GONDEREN adresi."""
    return (_kiraci_kayit().get("eposta") or "").strip()


# --------------------------------------------------------------------------- #
# Cok-mukellef baglam (1 kiraci = N mukellef). Aktif mukellef ?musteri=KOD ile.
# --------------------------------------------------------------------------- #
_CTX = threading.local()


# Kiracida mukellefler.json yoksa kullanilan hazir demo konfig — paketlenmis
# ornek ekstrelerle (ornek_veri/) calisir; yeni kiraci aninda demo yapabilir.
VARSAYILAN_MUKELLEFLER = {
    "_aciklama": "Varsayilan demo mukellefleri (kiraciya ozel mukellefler.json yazilana kadar).",
    "varsayilan": "MUK-001",
    "mukellefler": [
        {"kod": "MUK-001", "unvan": "Demir Çelik Sanayi A.Ş.", "sektor": "Metal / Üretim",
         "donem": "Mayıs 2026", "m": 18, "i": 1, "b": 1, "motor": True,
         "veri": {"bizim": "ornek/bizim_ekstreler.xlsx",
                  "karsi": "ornek/karsi_ekstreler.xlsx",
                  "gib": "ornek/gib_efatura_kayitlari.xlsx"}},
        {"kod": "MUK-002", "unvan": "Anadolu Tekstil San. Tic.", "sektor": "Tekstil",
         "donem": "Mayıs 2026", "m": 24, "i": 0, "b": 0, "motor": False},
        {"kod": "MUK-003", "unvan": "Gül Plastik Ltd. Şti.", "sektor": "Plastik / Ambalaj",
         "donem": "Mayıs 2026", "m": 9, "i": 3, "b": 2, "motor": False},
        {"kod": "MUK-015", "unvan": "Gaziantep Baharat Ltd.", "sektor": "Gıda / Baharat",
         "donem": "Mayıs 2026", "m": 0, "i": 0, "b": 3, "motor": True,
         "veri": {"bizim": "ornek/muk015_bizim.xlsx",
                  "karsi": "ornek/muk015_karsi.xlsx",
                  "gib": "ornek/muk015_gib.xlsx"}},
    ],
}


def mukellefler_yukle():
    cfg = _json_oku(MUKELLEFLER(), None)
    if cfg and cfg.get("mukellefler"):
        return cfg
    return VARSAYILAN_MUKELLEFLER


def mukellef_listesi():
    return mukellefler_yukle().get("mukellefler", [])


def _mk_yol(mk, anahtar):
    return _yol_coz((mk.get("veri") or {}).get(anahtar), anahtar)


def _canli_sayim(mk):
    s = G.kontrol(_mk_yol(mk, "bizim"), _mk_yol(mk, "karsi"), _mk_yol(mk, "gib"))
    mut = sum(1 for d in s.values() if not d["bulgular"])
    return mut, len(s) - mut, 0


def mukellef_listesi_canli():
    tum = _durum_tum()
    out = []
    for mk in mukellef_listesi():
        e = dict(mk)
        if mk.get("motor"):
            try:
                ds = tum.get(mk.get("kod"), {})
                g = ds.get(GKEY, {}) if isinstance(ds, dict) else {}
                baslandi = bool(g.get("tur", 0)) or any(
                    isinstance(v, dict) and v.get("gonderim", 0)
                    for k, v in ds.items() if k != GKEY)
                e["baslandi"] = baslandi
                cm, ci, cb = _canli_sayim(mk)
                if baslandi:
                    e["m"], e["i"], e["b"] = cm, ci, cb
                else:
                    e["m"], e["i"], e["b"] = 0, 0, cm + ci + cb
                e["canli"] = True
            except Exception:
                pass
        out.append(e)
    return out


def varsayilan_musteri():
    cfg = mukellefler_yukle()
    if cfg.get("varsayilan"):
        return cfg["varsayilan"]
    liste = cfg.get("mukellefler", [])
    return liste[0]["kod"] if liste else "MUK-001"


def musteri_coz(kod=None):
    liste = mukellef_listesi()
    if not liste:
        return {"kod": "MUK-001", "unvan": FIRMA, "donem": DONEM, "motor": True, "veri": {}}
    kod = kod or varsayilan_musteri()
    for mk in liste:
        if mk.get("kod") == kod:
            return mk
    return liste[0]


def set_aktif(kod):
    _CTX.musteri = musteri_coz(kod)


def aktif():
    mk = getattr(_CTX, "musteri", None)
    return mk if mk else musteri_coz(None)


def aktif_kod():
    return aktif().get("kod", "MUK-001")


def aktif_yol(anahtar):
    return _yol_coz((aktif().get("veri") or {}).get(anahtar), anahtar)


def aktif_unvan():
    return aktif().get("unvan", FIRMA)


def aktif_donem():
    return aktif().get("donem", DONEM)


# --------------------------------------------------------------------------- #
# Durum deposu — mukellef bazinda izole
# --------------------------------------------------------------------------- #
def _durum_ham():
    yol = DEMO_DURUM()
    if os.path.exists(yol):
        with open(yol, encoding="utf-8") as f:
            return json.load(f)
    return {}


def _eski_format_mi(ham):
    if not ham:
        return False
    if GKEY in ham:
        return True
    return not all(str(k).startswith("MUK-") for k in ham)


def _durum_tum():
    ham = _durum_ham()
    if _eski_format_mi(ham):
        return {varsayilan_musteri(): ham}
    return ham


def durum_yukle(musteri=None):
    return _durum_tum().get(musteri or aktif_kod(), {})


def durum_kaydet(d, musteri=None):
    # Kilit + gecici dosya + os.replace: es zamanli iki yanit/yazim durum
    # dosyasini yarim birakamaz (depo._yaz ile ayni desen).
    with _YAZ_KILIT:
        tum = _durum_tum()
        tum[musteri or aktif_kod()] = d
        yol = DEMO_DURUM()
        gecici = yol + ".tmp"
        with open(gecici, "w", encoding="utf-8") as f:
            json.dump(tum, f, ensure_ascii=False, indent=2)
        os.replace(gecici, yol)


# --------------------------------------------------------------------------- #
# Kanit zinciri (append-only JSONL, sha256 hash zinciri)
# --------------------------------------------------------------------------- #
def _son_hash():
    yol = KANIT_LOG()
    if not os.path.exists(yol):
        return "0" * 64
    son = "0" * 64
    with open(yol, encoding="utf-8") as f:
        for ln in f:
            ln = ln.strip()
            if ln:
                try:
                    son = json.loads(ln).get("hash", son)
                except Exception:
                    pass
    return son


def kanit_yaz(kayit):
    # Kilit: son-hash okuma + ekleme tek atomik adim olsun (zincir kopmasin).
    with _YAZ_KILIT:
        onceki = _son_hash()
        satir = {"zaman": datetime.now().isoformat(timespec="seconds"), **kayit,
                 "onceki_hash": onceki}
        ozet_girdi = json.dumps(satir, ensure_ascii=False, sort_keys=True)
        satir["hash"] = hashlib.sha256(ozet_girdi.encode("utf-8")).hexdigest()
        with open(KANIT_LOG(), "a", encoding="utf-8") as f:
            f.write(json.dumps(satir, ensure_ascii=False) + "\n")
        return satir


def kanit_dogrula():
    yol = KANIT_LOG()
    if not os.path.exists(yol):
        return {"satir": 0, "saglam": True, "bozuk": []}
    bozuk, onceki, i = [], "0" * 64, 0
    with open(yol, encoding="utf-8") as f:
        for ln in f:
            ln = ln.strip()
            if not ln:
                continue
            i += 1
            r = json.loads(ln)
            beklenen = r.get("hash")
            gecici = {k: v for k, v in r.items() if k != "hash"}
            gecici["onceki_hash"] = onceki
            h = hashlib.sha256(json.dumps(gecici, ensure_ascii=False, sort_keys=True).encode()).hexdigest()
            if h != beklenen or r.get("onceki_hash") != onceki:
                bozuk.append(i)
            onceki = beklenen
    return {"satir": i, "saglam": len(bozuk) == 0, "bozuk": bozuk}


# --------------------------------------------------------------------------- #
# E-posta (Gmail SMTP) — varsayilan KAPALI
# --------------------------------------------------------------------------- #
def cari_mail(ck):
    cfg = _json_oku(CARI_MAIL(), {})
    cariler = cfg.get("cariler", {})
    return cariler.get(ck) or cfg.get("varsayilan_karsi") or ""


def mail_ayar():
    return _json_oku(MAIL_AYAR(), {})


def mail_aktif_mi():
    return bool((mail_ayar().get("smtp_app_sifre") or "").strip())


def mail_gonder(alici, konu, html, metin=""):
    a = mail_ayar()
    sifre = (a.get("smtp_app_sifre") or "").strip().replace(" ", "")
    user = (a.get("smtp_user") or kiraci_mail())   # ayarda yoksa kiraci giris e-postasi
    if not sifre or not user or not alici:
        return False, "yapilandirma eksik (app sifresi/alici)", ""
    msg = EmailMessage()
    cfg = _json_oku(CARI_MAIL(), {})
    gad = cfg.get("_gonderici_ad") or aktif_unvan()
    mid = make_msgid(domain="mutabakat.local")
    msg["Message-ID"] = mid
    msg["From"] = f"{gad} <{user}>"
    msg["To"] = alici
    msg["Subject"] = konu
    msg.set_content(metin or "Bu e-posta HTML formatindadir.")
    msg.add_alternative(html, subtype="html")
    try:
        ctx = ssl.create_default_context()
        with smtplib.SMTP_SSL(a.get("smtp_host", "smtp.gmail.com"),
                              int(a.get("smtp_port", 465)), context=ctx, timeout=20) as s:
            s.login(user, sifre)
            s.send_message(msg)
        return True, "", mid
    except Exception as e:
        return False, str(e), mid


def _mail_govde(ck, adi, bakiye):
    # Link imzali token tasir: karsi taraf oturumsuz acar, token (kiraci, cari,
    # mukellef) uclusune kilitlidir (bkz. portal_token / app.py public kapisi).
    link = (f"{PORTAL_URL()}/portal?cari={ck}&musteri={aktif_kod()}"
            f"&tok={portal_token(ck)}")
    bak = f"{bakiye:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".") if bakiye is not None else "-"
    return f"""<div style="font-family:Segoe UI,Arial,sans-serif;max-width:560px;margin:auto;
      color:#1a2230;border:1px solid #e2e8f0;border-radius:14px;overflow:hidden">
      <div style="background:#2563eb;color:#fff;padding:18px 22px">
        <div style="font-size:12px;letter-spacing:1px;opacity:.85">MUTABAKAT FORMU · {aktif_donem()}</div>
        <div style="font-size:19px;font-weight:800;margin-top:4px">{aktif_unvan()}</div>
      </div>
      <div style="padding:22px">
        <p>Sayın <b>{adi}</b>,</p>
        <p>{aktif_donem()} dönemi cari hesap mutabakatı için kayıtlarımızdaki bakiyeniz:</p>
        <div style="background:#eaf1ff;border-left:4px solid #2563eb;border-radius:10px;
          padding:14px 18px;font-size:20px;font-weight:800;margin:14px 0">{bak} TL</div>
        <p>Mutabık olup olmadığınızı aşağıdaki butondan bildirebilir, mutabık değilseniz
          kendi ekstrenizi yükleyebilirsiniz. Sistem farkları iki taraf için de gösterir.</p>
        <p style="text-align:center;margin:24px 0">
          <a href="{link}" style="background:#2563eb;color:#fff;text-decoration:none;
            padding:13px 28px;border-radius:10px;font-weight:700;display:inline-block">
            Mutabakat Formunu Aç</a></p>
        <p style="font-size:12px;color:#5b6878">Buton çalışmazsa: <a href="{link}">{link}</a></p>
        <p style="font-size:12px;color:#5b6878;border-top:1px solid #e2e8f0;padding-top:12px">
          Cari kodu: {ck} · Bu e-posta Akıllı Mutabakat sistemi tarafından gönderilmiştir.</p>
      </div>
    </div>"""


# --------------------------------------------------------------------------- #
# OTP
# --------------------------------------------------------------------------- #
def _maskele_mail(mail):
    if not mail or "@" not in mail:
        return mail or ""
    ad, alan = mail.split("@", 1)
    if len(ad) <= 2:
        gizli = ad[0] + "*"
    else:
        gizli = ad[0] + "*" * (len(ad) - 2) + ad[-1]
    return f"{gizli}@{alan}"


def _otp_govde(ck, kod):
    return f"""<div style="font-family:Segoe UI,Arial,sans-serif;max-width:480px;margin:auto;
      color:#1a2230;border:1px solid #e2e8f0;border-radius:14px;overflow:hidden">
      <div style="background:#0f766e;color:#fff;padding:16px 22px">
        <div style="font-size:12px;letter-spacing:1px;opacity:.85">KİMLİK DOĞRULAMA KODU</div>
        <div style="font-size:18px;font-weight:800;margin-top:4px">{aktif_unvan()}</div>
      </div>
      <div style="padding:22px;text-align:center">
        <p style="text-align:left">Mutabakat yanıtınızı onaylamak için aşağıdaki tek kullanımlık
          kodu portala girin:</p>
        <div style="font-size:34px;font-weight:900;letter-spacing:10px;color:#0f766e;
          background:#e6fffb;border-radius:12px;padding:16px;margin:14px 0">{kod}</div>
        <p style="font-size:12px;color:#5b6878;text-align:left">Kod {OTP_GECERLILIK} dakika geçerlidir.</p>
        <p style="font-size:12px;color:#5b6878;text-align:left;border-top:1px solid #e2e8f0;
          padding-top:10px">Cari kodu: {ck} · Akıllı Mutabakat sistemi</p>
      </div>
    </div>"""


def _otp_anahtar(ck):
    """OTP anahtari kiraci+mukellef+cari uclusudur: iki kiracida ayni cari kodu
    (or. '120.01') olmasi normaldir; kodlar birbirinin uzerine yazilmamali."""
    return f"{depo.aktif_kiraci()}|{aktif_kod()}|{ck}"


def otp_iste(ck):
    kod = f"{random.randint(0, 999999):06d}"
    OTP_DEPO[_otp_anahtar(ck)] = {"kod": kod, "bitis": datetime.now() + timedelta(minutes=OTP_GECERLILIK)}
    alici = cari_mail(ck)
    if mail_aktif_mi():
        konu = f"Doğrulama Kodu · {aktif_donem()} · {aktif_unvan()}"
        ok, hata, mid = mail_gonder(alici, konu, _otp_govde(ck, kod))
        kanit_yaz({"olay": "OTP_GONDERILDI", "cari": ck, "alici": alici,
                   "kanal": "e-posta", "message_id": mid,
                   "sonuc": "gonderildi" if ok else f"hata: {hata}"})
        return {"ok": ok, "kanal": "e-posta", "maskeli_mail": _maskele_mail(alici),
                "hata": "" if ok else hata}
    kanit_yaz({"olay": "OTP_GONDERILDI", "cari": ck, "alici": alici,
               "kanal": "demo", "sonuc": "demo_kod_ekranda"})
    return {"ok": True, "kanal": "demo", "maskeli_mail": _maskele_mail(alici), "demo_kod": kod}


def otp_dogrula(ck, kod):
    anahtar = _otp_anahtar(ck)
    kayit = OTP_DEPO.get(anahtar)
    if not kayit:
        return False, "kod istenmedi veya suresi doldu"
    if datetime.now() > kayit["bitis"]:
        OTP_DEPO.pop(anahtar, None)
        return False, "kodun suresi doldu"
    if (kod or "").strip() != kayit["kod"]:
        return False, "kod hatali"
    OTP_DEPO.pop(anahtar, None)
    return True, ""


# --------------------------------------------------------------------------- #
# Karsi taraf PORTAL erisimi — oturumsuz, IMZALI baglanti (token)
#
# Maildeki linke tiklayan karsi tarafin platform oturumu yoktur; portal ve
# portala hizmet eden dar API kumesi, linkteki HMAC imzali token ile acilir.
# Token (kid|cari|musteri) uclusune baglidir: baska kiraci/cari/mukellef icin
# kullanilamaz. Anahtar kasa anahtarindan turetilir (KASA_ANAHTARI env'i).
# Token mutabakat donemi boyunca gecerlidir (mail linki gunlerce sonra
# acilabilir); yanitin kendisi ayrica OTP ile dogrulanir.
# --------------------------------------------------------------------------- #
PORTAL_GET_ACIK = {"/portal", "/portal.html", "/api/form", "/api/maildurum",
                   "/api/itiraz", "/api/cari_fark"}
PORTAL_POST_ACIK = {"/api/otp_iste", "/api/yanit"}


def _portal_imza(kid, ck, musteri):
    mesaj = f"{kid}|{ck}|{musteri}".encode("utf-8")
    return hmac.new(kasa._anahtar(), mesaj, hashlib.sha256).hexdigest()[:32]


def portal_token(ck, musteri=None, kid=None):
    """Karsi tarafa gonderilecek linkin imzasi. kid varsayilani aktif kiraci."""
    kid = kid or depo.aktif_kiraci()
    musteri = musteri if musteri is not None else aktif_kod()
    return f"{kid}.{_portal_imza(kid, ck, musteri)}"


def portal_coz(sub, params, post=False):
    """Public istek dogrulama. params: GET'te query dict (list degerli),
    POST'ta body dict. Gecerliyse {"kid","cari","musteri"} doner; degilse None."""
    izinli = PORTAL_POST_ACIK if post else PORTAL_GET_ACIK
    if (sub or "/") not in izinli:
        return None

    def _al(ad):
        v = params.get(ad)
        if isinstance(v, list):
            v = v[0] if v else ""
        return (v or "").strip()

    tok, ck, musteri = _al("tok"), _al("cari"), _al("musteri")
    if not tok or not ck or "." not in tok:
        return None
    kid, _, imza = tok.partition(".")
    if not kid or not kiraci.kiraci_getir(kid):
        return None
    beklenen = _portal_imza(kid, ck, musteri)
    if not hmac.compare_digest(imza, beklenen):
        return None
    return {"kid": kid, "cari": ck, "musteri": musteri}


# --------------------------------------------------------------------------- #
# Asama veri fonksiyonlari
# --------------------------------------------------------------------------- #
def asama_yukle():
    b, k, g = M.oku(aktif_yol("bizim")), M.oku(aktif_yol("karsi")), M.oku(aktif_yol("gib"))
    def prev(rows):
        return [{"cari": r["cari_kodu"], "tarih": M.tarih_str(r["tarih"]),
                 "belge_no": r["belge_no"], "belge_tipi": r.get("belge_tipi"),
                 "tutar": float(r["tutar"]), "tip": r.get("tip", "")}
                for r in rows[:8]]
    tipler = {}
    for r in b + k:
        t = str(r.get("belge_tipi"))
        tipler[t] = tipler.get(t, 0) + 1
    return {"bizim": len(b), "karsi": len(k), "gib": len(g),
            "tipler": tipler, "onizleme": prev(b)}


def asama_gib():
    s = G.kontrol(aktif_yol("bizim"), aktif_yol("karsi"), aktif_yol("gib"))
    bize, karsiya, temiz = [], [], []
    for ck, d in s.items():
        gibsel = [(sa, ke, ac) for (asama, sa, ke, ac) in d["bulgular"] if ke != "EKSTRE"]
        if not gibsel:
            continue
        for sa, ke, ac in gibsel:
            item = {"cari": ck, "adi": d["adi"], "kesinlik": ke, "ac": ac}
            (bize if sa == "BIZ" else karsiya).append(item)
    for ck, d in s.items():
        if not d["bulgular"]:
            temiz.append(f"{ck} {d['adi']}")
    return {"bize": bize, "karsiya": karsiya, "temiz": temiz}


def _nonef(rows):
    return [dict(r) for r in rows
            if not (str(r.get("belge_tipi")) == "e-Fatura" and r.get("tip") == "FATURA")]


def asama_eslestir():
    bizim = M.grupla(M.oku(aktif_yol("bizim")))
    karsi = M.grupla(M.oku(aktif_yol("karsi")))
    sayac = {"EXACT": 0, "FUZZY": 0, "SUBSET_SUM": 0}
    ornekler = []
    for ck in sorted(set(bizim) | set(karsi)):
        for tip, b, k, ac in M.cari_esle(_nonef(bizim.get(ck, [])), _nonef(karsi.get(ck, []))):
            if tip in sayac:
                sayac[tip] += 1
                if tip in ("FUZZY", "SUBSET_SUM"):
                    ornekler.append({"cari": ck, "tip": tip, "ac": ac})
    return {"sayac": sayac, "ornekler": ornekler}


def asama_triyaj():
    s = G.kontrol(aktif_yol("bizim"), aktif_yol("karsi"), aktif_yol("gib"))
    analiz = D.cari_analiz(aktif_yol("bizim"), aktif_yol("karsi"))
    cariler, mutabik = [], 0
    for ck, d in s.items():
        durum = "MUTABIK" if not d["bulgular"] else "AKSIYON"
        if durum == "MUTABIK":
            mutabik += 1
        a = analiz.get(ck, {})
        cariler.append({"cari": ck, "adi": d["adi"], "durum": durum,
                        "bizim_bakiye": a.get("bizim_bakiye", 0),
                        "karsi_bakiye": a.get("karsi_bakiye", 0),
                        "fark_sayisi": len(d["bulgular"])})
    cariler.sort(key=lambda c: (c["durum"] == "MUTABIK", c["cari"]))
    return {"toplam": len(s), "mutabik": mutabik,
            "aksiyon": len(s) - mutabik, "cariler": cariler}


def asama_form():
    analiz = D.cari_analiz(aktif_yol("bizim"), aktif_yol("karsi"))
    durum = durum_yukle()
    out = []
    for ck, a in analiz.items():
        esit = abs(a["bizim_bakiye"] - a["karsi_bakiye"]) <= 0.01
        out.append({"cari": ck, "adi": a["cari_adi"],
                    "bizim_bakiye": a["bizim_bakiye"], "karsi_bakiye": a["karsi_bakiye"],
                    "esit": esit, "motor_durum": a["motor_durum"],
                    "yanit": durum.get(ck, {})})
    return {"firma": aktif_unvan(), "donem": aktif_donem(), "musteri": aktif_kod(), "cariler": out}


def asama_takip():
    analiz = D.cari_analiz(aktif_yol("bizim"), aktif_yol("karsi"))
    durum = durum_yukle()
    g = durum.get(GKEY, {})
    rows = []
    sayac = {"BEKLIYOR": 0, "MUTABIK": 0, "ITIRAZLI": 0}
    for ck, a in analiz.items():
        y = durum.get(ck, {})
        d = y.get("durum", "BEKLIYOR")
        sayac[d] = sayac.get(d, 0) + 1
        rows.append({"cari": ck, "adi": a["cari_adi"],
                     "bizim_bakiye": a["bizim_bakiye"], "karsi_bakiye": a["karsi_bakiye"],
                     "durum": d, "dosya": y.get("dosya"), "not": y.get("not", ""),
                     "gonderim": y.get("gonderim", 0), "son_gonderim": y.get("son_gonderim")})
    return {"sayac": sayac, "cariler": rows,
            "tur": g.get("tur", 0), "log": g.get("log", [])}


def gonder_tur(secili=None):
    analiz = D.cari_analiz(aktif_yol("bizim"), aktif_yol("karsi"))
    d = durum_yukle()
    g = d.get(GKEY, {"tur": 0, "log": []})
    g["tur"] += 1
    tur = g["tur"]
    simdi = datetime.now().strftime("%d.%m.%Y %H:%M")
    secset = set(secili) if secili else None
    gonderilen, atlanan = [], []
    for ck, a in analiz.items():
        if secset is not None and ck not in secset:
            continue
        y = d.get(ck, {})
        if y.get("durum") in ("MUTABIK", "ITIRAZLI"):
            atlanan.append({"cari": ck, "adi": a["cari_adi"], "durum": y["durum"]})
            continue
        y["gonderim"] = y.get("gonderim", 0) + 1
        y["son_gonderim"] = simdi
        alici = cari_mail(ck)
        mail_ok, mail_hata, mid = False, "kapali", ""
        if mail_aktif_mi():
            konu = f"Mutabakat Formu · {aktif_donem()} · {aktif_unvan()}"
            html = _mail_govde(ck, a["cari_adi"], a.get("bizim_bakiye", 0))
            mail_ok, mail_hata, mid = mail_gonder(alici, konu, html)
            y["mail_alici"] = alici
            y["mail_son"] = "gonderildi" if mail_ok else f"hata: {mail_hata}"
            kanit_yaz({"olay": "FORM_GONDERILDI", "cari": ck, "adi": a["cari_adi"],
                       "alici": alici, "tur": tur, "kacinci": y["gonderim"],
                       "message_id": mid, "sonuc": "gonderildi" if mail_ok else f"hata: {mail_hata}"})
        else:
            # Demo modunda da gonderim KANITA islenir — kanit zinciri eksiksiz kalsin.
            kanit_yaz({"olay": "FORM_GONDERILDI", "cari": ck, "adi": a["cari_adi"],
                       "alici": alici, "tur": tur, "kacinci": y["gonderim"],
                       "kanal": "demo", "sonuc": "demo_gonderim (mail kapali)"})
        d[ck] = y
        gonderilen.append({"cari": ck, "adi": a["cari_adi"], "kacinci": y["gonderim"],
                           "mail": alici, "mail_ok": mail_ok,
                           "mail_hata": "" if mail_ok else mail_hata})
    g["log"].append({"tur": tur, "tarih": simdi, "tip": "secili" if secset else "toplu",
                     "gonderilen": len(gonderilen), "atlanan": len(atlanan)})
    d[GKEY] = g
    durum_kaydet(d)
    mail_ok_say = sum(1 for x in gonderilen if x.get("mail_ok"))
    return {"tur": tur, "tip": "secili" if secset else "toplu",
            "gonderilen": gonderilen, "atlanan": atlanan,
            "gonderilen_sayi": len(gonderilen), "atlanan_sayi": len(atlanan),
            "mail_aktif": mail_aktif_mi(), "mail_ok_sayi": mail_ok_say}


def kaydet_yanit(body, ip="", ua=""):
    ck = body.get("cari")
    secim = body.get("secim")        # MUTABIK | ITIRAZLI
    if mail_aktif_mi():
        ok, hata = otp_dogrula(ck, body.get("otp", ""))
        if not ok:
            return {"ok": False, "hata": hata, "otp_gerekli": True}
        kanit_yaz({"olay": "OTP_DOGRULANDI", "cari": ck, "ip": ip, "tarayici": ua})
    d = durum_yukle()
    kayit = d.get(ck, {})
    kayit.update({"durum": secim, "not": body.get("not", ""), "dosya": kayit.get("dosya")})
    b64 = body.get("dosya_b64")
    ad = body.get("dosya_adi")
    if b64 and ad:
        if "," in b64:
            b64 = b64.split(",", 1)[1]
        guvenli = f"{ck}__{os.path.basename(ad)}".replace(" ", "_")
        with open(os.path.join(_uploads(), guvenli), "wb") as f:
            f.write(base64.b64decode(b64))
        kayit["dosya"] = guvenli
    k = kanit_yaz({"olay": "YANIT_ALINDI", "cari": ck, "durum": secim,
                   "ip": ip, "tarayici": ua, "dosya": kayit.get("dosya"),
                   "not": body.get("not", "")})
    kayit["yanit_kanit"] = {"zaman": k["zaman"], "ip": ip}
    d[ck] = kayit
    durum_kaydet(d)
    return {"ok": True, "cari": ck, "durum": secim, "dosya": kayit["dosya"],
            "kanit_zaman": k["zaman"]}


def _parse_yuklenen(yol, ck, adi):
    from openpyxl import load_workbook
    wb = load_workbook(yol, read_only=True, data_only=True)
    ws = wb.active
    allrows = list(ws.iter_rows(values_only=True))
    hdr, hi = None, None
    for i, r in enumerate(allrows):
        cells = [str(c).strip().lower() if c is not None else "" for c in r]
        if any(c.startswith("tarih") for c in cells) and any("tutar" in c for c in cells):
            hdr, hi = cells, i
            break
    if hdr is None:
        return None

    def idx(pred):
        for j, c in enumerate(hdr):
            if pred(c):
                return j
        return None
    i_tar = idx(lambda c: "tarih" in c)
    i_bno = idx(lambda c: "belge" in c and ("no" in c or "evrak" in c))
    i_bti = idx(lambda c: "belge" in c and "tip" in c)
    i_tip = idx(lambda c: c == "tip" or ("tip" in c and "belge" not in c))
    i_ac = idx(lambda c: "acikla" in c or "açıkla" in c)
    i_tut = idx(lambda c: "tutar" in c)

    def hucre(r, j):
        return r[j] if (j is not None and j < len(r) and r[j] is not None) else None

    out = []
    sira = 0
    for r in allrows[hi + 1:]:
        if r is None or all(c is None for c in r):
            continue
        tut = hucre(r, i_tut)
        if tut in (None, ""):
            continue
        try:
            tut = float(tut)
        except (TypeError, ValueError):
            continue
        t = hucre(r, i_tar)
        if isinstance(t, str):
            t2 = None
            for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y"):
                try:
                    t2 = datetime.strptime(t.strip(), fmt).date()
                    break
                except ValueError:
                    pass
            t = t2 or datetime(2026, 5, 1).date()
        elif hasattr(t, "date"):
            t = t.date()
        elif t is None:
            t = datetime(2026, 5, 1).date()
        out.append({
            "cari_kodu": ck, "cari_adi": adi, "tarih": t,
            "belge_no": str(hucre(r, i_bno) or ""),
            "belge_tipi": str(hucre(r, i_bti) or ""),
            "aciklama": str(hucre(r, i_ac) or ""),
            "tutar": round(tut, 2),
            "tip": str(hucre(r, i_tip) or "FATURA").upper(),
            "_id": sira, "_eslesti": False,
        })
        sira += 1
    return out


def _kid(prefix, s):
    return prefix + hashlib.sha1(str(s).encode("utf-8")).hexdigest()[:10]


def _fark_listele(bizim, karsi, gib):
    biz_eksik, karsi_eksik, tutar_farki = [], [], []
    b_ef = G._efatura_listesi(bizim)
    k_ef = G._efatura_listesi(karsi)
    b_kull, k_kull = set(), set()
    for g in gib:
        tut = float(g["tutar"])
        b = G._esle_bul(g, b_ef, b_kull)
        k = G._esle_bul(g, k_ef, k_kull)
        if b is None:
            biz_eksik.append({"id": _kid("GBE:", g["belge_no"]), "kaynak": "GİB",
                              "belge": g["belge_no"], "tutar": tut,
                              "ac": f"GİB'de var, bizim defterde yok ({g['belge_no']} = {tut:,.2f} TL)"})
        elif not M.yakin(b["tutar"], tut):
            tutar_farki.append({"id": _kid("GFB:", g["belge_no"]), "taraf": "BIZ", "belge": g["belge_no"],
                                "ac": f"{g['belge_no']}: bizde {float(b['tutar']):,.2f} / GİB {tut:,.2f} → bizim kayıt hatalı"})
        if k is None:
            karsi_eksik.append({"id": _kid("GKE:", g["belge_no"]), "kaynak": "GİB",
                                "belge": g["belge_no"], "tutar": tut,
                                "ac": f"GİB'de var, karşının yüklediği ekstrede yok ({g['belge_no']} = {tut:,.2f} TL)"})
        elif not M.yakin(k["tutar"], tut):
            tutar_farki.append({"id": _kid("GFK:", g["belge_no"]), "taraf": "KARSI", "belge": g["belge_no"],
                                "ac": f"{g['belge_no']}: karşıda {float(k['tutar']):,.2f} / GİB {tut:,.2f} → karşı kayıt hatalı"})
    for tip, bb, kk, ac in M.cari_esle(_nonef(bizim), _nonef(karsi)):
        if tip == "EKSIK_KARSIDA":
            karsi_eksik.append({"id": _kid("EKE:", ac), "kaynak": "ekstre", "ac": ac})
        elif tip == "EKSIK_BIZDE":
            biz_eksik.append({"id": _kid("EBE:", ac), "kaynak": "ekstre", "ac": ac})
        elif tip == "TUTAR_FARKI":
            tutar_farki.append({"id": _kid("EF:", ac), "taraf": "BELIRSIZ", "ac": ac})
    return biz_eksik, karsi_eksik, tutar_farki


def _filtre_manuel(ck, biz_eksik, karsi_eksik, tutar_farki):
    me = (durum_yukle().get(ck, {}) or {}).get("manuel_eslesme", {}) or {}
    if not me:
        return biz_eksik, karsi_eksik, tutar_farki, []
    tum = {x["id"]: x for x in (biz_eksik + karsi_eksik + tutar_farki)}
    def kalan(lst):
        return [x for x in lst if x["id"] not in me]
    manuel = []
    for mid, kayit in me.items():
        x = tum.get(mid)
        manuel.append({"id": mid, "ac": (x or {}).get("ac", kayit.get("ac", mid)),
                       "tip": kayit.get("tip", "kapat"), "es": kayit.get("es"),
                       "not": kayit.get("not", ""), "zaman": kayit.get("zaman", ""),
                       "kim": kayit.get("kim", "")})
    return kalan(biz_eksik), kalan(karsi_eksik), kalan(tutar_farki), manuel


def itiraz_analiz(ck):
    durum = durum_yukle()
    y = durum.get(ck, {})
    dosya = y.get("dosya")
    if not dosya:
        return {"parse": False, "mesaj": "Bu cari icin yuklenmis ekstre yok."}
    yol = os.path.join(_uploads(), dosya)
    if not dosya.lower().endswith((".xlsx", ".xlsm")):
        return {"parse": False, "dosya": dosya,
                "mesaj": "Yuklenen dosya Excel degil (PDF/foto). Otomatik ayristirma "
                         "yalnizca Excel icin; belge saklandi, elle incelenebilir."}

    bizim = M.grupla(M.oku(aktif_yol("bizim"))).get(ck, [])
    gib = M.grupla(M.oku(aktif_yol("gib"))).get(ck, [])
    adi = (bizim or [{"cari_adi": ck}])[0]["cari_adi"]
    karsi = _parse_yuklenen(yol, ck, adi)
    if karsi is None:
        return {"parse": False, "dosya": dosya,
                "mesaj": "Excel basligi taninamadi (Tarih/Tutar sutunlari bulunamadi)."}

    biz_eksik, karsi_eksik, tutar_farki = _fark_listele(bizim, karsi, gib)
    biz_eksik, karsi_eksik, tutar_farki, manuel = _filtre_manuel(ck, biz_eksik, karsi_eksik, tutar_farki)

    def _bak(rows):
        fat = sum(float(r["tutar"]) for r in rows if r.get("tip") == "FATURA")
        ode = sum(float(r["tutar"]) for r in rows if r.get("tip") == "ODEME")
        return round(fat - ode, 2)
    bizim_bakiye = _bak(bizim)
    karsi_bakiye = _bak(karsi)

    return {"parse": True, "dosya": dosya, "cari": ck, "adi": adi,
            "karsi_satir": len(karsi),
            "bizim_bakiye": bizim_bakiye, "karsi_bakiye": karsi_bakiye,
            "fark": round(abs(bizim_bakiye - karsi_bakiye), 2),
            "biz_eksik": biz_eksik, "karsi_eksik": karsi_eksik, "tutar_farki": tutar_farki,
            "manuel": manuel,
            "mutabik": not (biz_eksik or karsi_eksik or tutar_farki)}


DEVIR_ANAHTAR = ("devir", "acil", "açıl", "devreden", "onceki donem", "önceki dönem",
                 "bakiye dev", "acilis", "açılış")


def _devir_mi(r):
    s = (str(r.get("belge_tipi", "")) + " " + str(r.get("aciklama", ""))).lower()
    return any(a in s for a in DEVIR_ANAHTAR)


def _imzali(r):
    t = float(r.get("tutar", 0) or 0)
    return t if r.get("tip") == "FATURA" else -t


def _devir_bul(b_rows, k_rows):
    bd = [r for r in b_rows if _devir_mi(r)]
    kd = [r for r in k_rows if _devir_mi(r)]
    if not bd and not kd:
        return {"var": False}
    biz = round(sum(_imzali(r) for r in bd), 2)
    karsi = round(sum(_imzali(r) for r in kd), 2)
    fark = round(biz - karsi, 2)
    if abs(fark) <= 0.01:
        ac = f"Devir/açılış bakiyesi mutabık ({biz:,.2f} TL)."
    else:
        ac = (f"DEVİR FARKI: bizdeki açılış {biz:,.2f} / karşıdaki {karsi:,.2f} "
              f"→ fark {fark:+,.2f} TL. Her dönem bu tutar kadar kayar; önce devir mutabakatı şart.")
    return {"var": True, "biz": biz, "karsi": karsi, "fark": fark,
            "mutabik": abs(fark) <= 0.01, "ac": ac}


def _capraz_bul(bizim_g, karsi_g):
    fazla, eksik = [], []
    for ck in sorted(set(bizim_g) | set(karsi_g)):
        b = [dict(r) for r in bizim_g.get(ck, [])]
        k = [dict(r) for r in karsi_g.get(ck, [])]
        adi = (b or k or [{"cari_adi": ck}])[0].get("cari_adi", ck)
        for tip, bx, kx, ac in M.cari_esle(b, k):
            if tip == "EKSIK_KARSIDA":
                fazla.append({"cari": ck, "adi": adi, "r": bx})
            elif tip == "EKSIK_BIZDE":
                eksik.append({"cari": ck, "adi": adi, "r": kx})
    adaylar = []
    for f in fazla:
        for e in eksik:
            if f["cari"] == e["cari"]:
                continue
            cf, ce = f["r"], e["r"]
            cek = M.sayi_cekirdek(cf["belge_no"])
            cek_ayni = bool(cek) and cek == M.sayi_cekirdek(ce["belge_no"])
            tut_ayni = M.yakin(float(cf["tutar"]), float(ce["tutar"]))
            if not (cek_ayni or tut_ayni):
                continue
            kesinlik = "YÜKSEK" if (cek_ayni and tut_ayni) else "ORTA"
            adaylar.append({
                "kesinlik": kesinlik,
                "fazla_cari": f["cari"], "fazla_adi": f["adi"],
                "eksik_cari": e["cari"], "eksik_adi": e["adi"],
                "belge": cf["belge_no"], "tutar": float(cf["tutar"]),
                "ac": (f"{float(cf['tutar']):,.2f} TL belge ({cf['belge_no']}) "
                       f"{f['cari']} {f['adi']} kartına işlenmiş; ama {e['cari']} {e['adi']} "
                       f"tarafında bu tutar eksik ({ce['belge_no']}). "
                       f"Muhtemelen yanlış cariye işlendi → {e['cari']}'ye taşınmalı.")})
    sira = {"YÜKSEK": 0, "ORTA": 1}
    adaylar.sort(key=lambda x: sira.get(x["kesinlik"], 9))
    return adaylar


def caprazcari_tara():
    return {"adaylar": _capraz_bul(M.grupla(M.oku(aktif_yol("bizim"))), M.grupla(M.oku(aktif_yol("karsi"))))}


def cari_fark(ck):
    bizim = M.grupla(M.oku(aktif_yol("bizim"))).get(ck, [])
    karsi = M.grupla(M.oku(aktif_yol("karsi"))).get(ck, [])
    gib = M.grupla(M.oku(aktif_yol("gib"))).get(ck, [])
    adi = (bizim or karsi or [{"cari_adi": ck}])[0].get("cari_adi", ck)
    biz_eksik, karsi_eksik, tutar_farki = _fark_listele(bizim, karsi, gib)
    biz_eksik, karsi_eksik, tutar_farki, manuel = _filtre_manuel(ck, biz_eksik, karsi_eksik, tutar_farki)
    return {"cari": ck, "adi": adi, "biz_eksik": biz_eksik, "karsi_eksik": karsi_eksik,
            "tutar_farki": tutar_farki, "manuel": manuel, "devir": _devir_bul(bizim, karsi),
            "mutabik": not (biz_eksik or karsi_eksik or tutar_farki)}


def manuel_esle(body, ip="", ua=""):
    ck = body.get("cari")
    if not ck:
        return {"ok": False, "hata": "cari yok"}
    d = durum_yukle()
    kayit = d.get(ck, {})
    me = kayit.get("manuel_eslesme", {})
    simdi = datetime.now().isoformat(timespec="seconds")
    notu = body.get("not", "")
    eklenen = []
    for mid in (body.get("kapat") or []):
        me[mid] = {"tip": "kapat", "not": notu, "zaman": simdi, "kim": ip}
        eklenen.append(mid)
    for cift in (body.get("cift") or []):
        if len(cift) == 2:
            a, b = cift
            me[a] = {"tip": "cift", "es": b, "not": notu, "zaman": simdi, "kim": ip}
            me[b] = {"tip": "cift", "es": a, "not": notu, "zaman": simdi, "kim": ip}
            eklenen += [a, b]
    kayit["manuel_eslesme"] = me
    d[ck] = kayit
    durum_kaydet(d)
    k = kanit_yaz({"olay": "MANUEL_KAPAT", "cari": ck, "kalemler": eklenen,
                   "not": notu, "ip": ip, "tarayici": ua})
    return {"ok": True, "cari": ck, "eklenen": eklenen, "kanit_zaman": k["zaman"]}


def manuel_geri(body, ip="", ua=""):
    ck = body.get("cari")
    mid = body.get("id")
    d = durum_yukle()
    kayit = d.get(ck, {})
    me = kayit.get("manuel_eslesme", {})
    silinen = []
    es = (me.get(mid) or {}).get("es")
    for x in [mid, es]:
        if x and x in me:
            me.pop(x); silinen.append(x)
    kayit["manuel_eslesme"] = me
    d[ck] = kayit
    durum_kaydet(d)
    k = kanit_yaz({"olay": "MANUEL_GERI", "cari": ck, "kalemler": silinen, "ip": ip, "tarayici": ua})
    return {"ok": True, "cari": ck, "silinen": silinen, "kanit_zaman": k["zaman"]}


ROUTES = {
    "/api/yukle": asama_yukle, "/api/gib": asama_gib, "/api/eslestir": asama_eslestir,
    "/api/triyaj": asama_triyaj, "/api/form": asama_form, "/api/takip": asama_takip,
}

STATIK = {"muhasebe_ofisi.html", "mutabakat_panosu.html", "canli_akis.html"}

_JSON = "application/json; charset=utf-8"
_HTML = "text/html; charset=utf-8"


def _jb(obj):
    return json.dumps(obj, ensure_ascii=False).encode("utf-8")


_TEMA_STIL = """<style id="mt-tema-stil">
html[data-tema="antrasit"]{--bg:#15181e;--yuzey:#1e2129;--yuzey2:#262a33;--cizgi:rgba(255,255,255,.07);--metin:#f4f4f5;--soluk:#a1a1aa;}
html[data-tema="antrasit"] body{background-color:#15181e;background-image:radial-gradient(at 0% 0%,rgba(20,184,166,.14) 0,transparent 55%),radial-gradient(at 100% 0%,rgba(16,185,129,.10) 0,transparent 55%),radial-gradient(at 50% 100%,rgba(245,158,11,.05) 0,transparent 55%);}
html[data-tema="acik"]{--bg:#eef2f7;--yuzey:#ffffff;--yuzey2:#f1f5f9;--cizgi:rgba(15,23,42,.12);--metin:#0f172a;--soluk:#475569;--mavi:#0d9488;--mor:#4f46e5;}
html[data-tema="acik"] body{background-color:#eef2f7;background-image:radial-gradient(at 0% 0%,rgba(13,148,136,.10) 0,transparent 55%),radial-gradient(at 100% 0%,rgba(16,185,129,.08) 0,transparent 55%),radial-gradient(at 50% 100%,rgba(245,158,11,.06) 0,transparent 55%);}
.mt-tema-btn{position:fixed;top:14px;right:14px;z-index:9999;background:var(--yuzey2);color:var(--metin);border:1px solid var(--cizgi);border-radius:10px;padding:8px 12px;font:inherit;font-size:13px;font-weight:600;cursor:pointer;display:flex;align-items:center;gap:7px;box-shadow:0 4px 16px rgba(0,0,0,.25);}
.mt-tema-btn:hover{background:var(--yuzey);}
</style>"""

_TEMA_BTN = '<button class="mt-tema-btn" id="mt-tema-btn" onclick="mtTema()" title="Tema (Lacivert / Antrasit / Açık)">\U0001f319 Lacivert</button>'

_TEMA_JS = """<script>
(function(){
 var T=[{k:'lacivert',a:'Lacivert',i:'\\u{1F319}'},{k:'antrasit',a:'Antrasit',i:'\\u25D0'},{k:'acik',a:'A\\u00e7\\u0131k',i:'\\u2600'}];
 function uygula(k){
   document.documentElement.setAttribute('data-tema',k);
   var t=T.filter(function(x){return x.k===k;})[0]||T[0];
   var b=document.getElementById('mt-tema-btn');
   if(b)b.innerHTML='<span style="font-size:14px">'+t.i+'</span> '+t.a;
 }
 var m='lacivert';try{m=localStorage.getItem('aykapanis_tema')||'lacivert';}catch(e){}
 document.documentElement.setAttribute('data-tema',m);
 window.mtTema=function(){
   var a=document.documentElement.getAttribute('data-tema'),i=0;
   for(var k=0;k<T.length;k++){if(T[k].k===a){i=k;break;}}
   var y=T[(i+1)%T.length].k;
   try{localStorage.setItem('aykapanis_tema',y);}catch(e){}
   uygula(y);
 };
 document.addEventListener('DOMContentLoaded',function(){uygula(document.documentElement.getAttribute('data-tema'));});
})();
</script>"""


def _sayfa(adi):
    """HTML sablonunu okur, dahili mutlak yollari /mutabakat prefiksiyle yeniden yazar,
    ve ofis panosu basligini aktif kiracinin adina gore enjekte eder."""
    with open(os.path.join(TPL, adi), encoding="utf-8") as f:
        s = f.read()
    s = s.replace("/api/", "/mutabakat/api/").replace("/uploads/", "/mutabakat/uploads/")
    if adi == "muhasebe_ofisi.html":
        ad = ofis_adi()
        bas = (ad[:1] or "M").upper()
        s = s.replace('<div class="logo">A+</div>', f'<div class="logo">{bas}</div>')
        s = s.replace("ARTI <span>Mali Müşavirlik</span> — Aylık Kapanış Panosu",
                      f"<span>{ad}</span> — Mutabakat Panosu")
    # Ay Kapanis ile ayni 3 tema (Lacivert/Antrasit/Acik) - tum sablonlara enjekte
    s = s.replace("</head>", _TEMA_STIL + _TEMA_JS + "</head>", 1)
    s = s.replace("</body>", _TEMA_BTN + "</body>", 1)
    return s.encode("utf-8")


# --------------------------------------------------------------------------- #
# Mail / portal ayarlari (yalniz OTURUMLU ofis kullanicisi; public kapidan
# erisilemez — PORTAL_*_ACIK kumelerinde yoktur).
# --------------------------------------------------------------------------- #
def api_mailayar_oku():
    a = mail_ayar()
    cfg = _json_oku(CARI_MAIL(), {})
    return {"ok": True,
            "smtp_user": (a.get("smtp_user") or "").strip() or kiraci_mail(),
            "app_sifre_kayitli": bool((a.get("smtp_app_sifre") or "").strip()),
            "portal_url": (a.get("portal_url") or "").strip(),
            "portal_efektif": PORTAL_URL(),
            "gonderici_ad": cfg.get("_gonderici_ad", "") or aktif_unvan(),
            "varsayilan_karsi": cfg.get("varsayilan_karsi", ""),
            "cariler": cfg.get("cariler", {}),
            "mail_aktif": mail_aktif_mi()}


def api_mailayar_yaz(body, ip="", ua=""):
    """SMTP + portal + varsayilan karsi mail ayarlarini yazar.
    App sifresi yalniz doluysa guncellenir (bos = mevcut korunur); yanitlarda
    sifre asla geri donmez, kanita da yazilmaz."""
    sifre = (body.get("smtp_app_sifre") or "").strip()
    with _YAZ_KILIT:
        a = mail_ayar()
        if "smtp_user" in body:
            a["smtp_user"] = (body.get("smtp_user") or "").strip()
        if sifre:
            a["smtp_app_sifre"] = sifre
        if body.get("sifre_sil"):
            a["smtp_app_sifre"] = ""
        if "portal_url" in body:
            a["portal_url"] = (body.get("portal_url") or "").strip()
        _json_yaz(MAIL_AYAR(), a)

        cfg = _json_oku(CARI_MAIL(), {})
        if "gonderici_ad" in body:
            cfg["_gonderici_ad"] = (body.get("gonderici_ad") or "").strip()
        if "varsayilan_karsi" in body:
            cfg["varsayilan_karsi"] = (body.get("varsayilan_karsi") or "").strip()
        cariler = body.get("cariler")
        if isinstance(cariler, dict):
            mevcut = cfg.get("cariler", {})
            for k, v in cariler.items():
                v = (v or "").strip()
                if v:
                    mevcut[str(k)] = v
                else:
                    mevcut.pop(str(k), None)
            cfg["cariler"] = mevcut
        _json_yaz(CARI_MAIL(), cfg)
    degisen = [k for k in ("smtp_user", "portal_url", "gonderici_ad",
                           "varsayilan_karsi", "cariler", "sifre_sil") if k in body]
    if sifre:
        degisen.append("smtp_app_sifre(guncellendi)")
    kanit_yaz({"olay": "MAIL_AYAR_GUNCELLENDI", "alanlar": degisen, "ip": ip, "tarayici": ua})
    return api_mailayar_oku()


# --------------------------------------------------------------------------- #
# Platform Handler'inin /mutabakat/* icin cagirdigi dispatch arayuzu.
# Donus: (kod:int, ctype:str, govde:bytes)
#
# public: app.py'nin imzali token ile dogruladigi oturumsuz karsi-taraf istegi
#         ({"kid","cari","musteri"}). Bu modda YALNIZ portal rotalari servis
#         edilir; cari/mukellef token'dakine SABITLENIR (parametre oynamasi
#         baska cariyi acamaz) ve maildurum yalniz aktiflik bayragi doner.
# --------------------------------------------------------------------------- #
def dispatch_get(subpath, query, ip="", host="", public=None):
    _CTX.host = host
    if public:
        set_aktif(public.get("musteri") or None)
    else:
        set_aktif((query.get("musteri") or [None])[0])
    p = subpath or "/"
    try:
        if public:
            if p not in PORTAL_GET_ACIK:
                return 404, _JSON, _jb({"hata": "bulunamadi"})
            pck = public.get("cari", "")
            if p in ("/portal", "/portal.html"):
                return 200, _HTML, _sayfa("karsi_taraf_portal.html")
            if p == "/api/form":
                d = asama_form()
                d["cariler"] = [c for c in d.get("cariler", []) if c.get("cari") == pck]
                return 200, _JSON, _jb(d)
            if p == "/api/maildurum":
                return 200, _JSON, _jb({"aktif": mail_aktif_mi()})
            if p == "/api/itiraz":
                return 200, _JSON, _jb(itiraz_analiz(pck))
            if p == "/api/cari_fark":
                return 200, _JSON, _jb(cari_fark(pck))
            return 404, _JSON, _jb({"hata": "bulunamadi"})

        if p in ("", "/", "/index.html"):
            return 200, _HTML, _sayfa("index.html")
        if p in ("/portal", "/portal.html"):
            return 200, _HTML, _sayfa("karsi_taraf_portal.html")
        if p == "/ofis":
            return 200, _HTML, _sayfa("muhasebe_ofisi.html")
        ad = os.path.basename(p)
        if ad in STATIK:
            return 200, _HTML, _sayfa(ad)
        if p == "/api/mukellefler":
            return 200, _JSON, _jb({"varsayilan": varsayilan_musteri(),
                                    "mukellefler": mukellef_listesi_canli()})
        if p in ROUTES:
            return 200, _JSON, _jb(ROUTES[p]())
        if p == "/api/itiraz":
            return 200, _JSON, _jb(itiraz_analiz((query.get("cari") or [""])[0]))
        if p == "/api/cari_fark":
            return 200, _JSON, _jb(cari_fark((query.get("cari") or [""])[0]))
        if p == "/api/caprazcari":
            return 200, _JSON, _jb(caprazcari_tara())
        if p == "/api/mailayar":
            return 200, _JSON, _jb(api_mailayar_oku())
        if p == "/api/maildurum":
            cfg = _json_oku(CARI_MAIL(), {})
            return 200, _JSON, _jb({"aktif": mail_aktif_mi(),
                                    "gonderici": cfg.get("_gonderici", ""),
                                    "varsayilan_karsi": cfg.get("varsayilan_karsi", ""),
                                    "cariler": cfg.get("cariler", {}),
                                    "cari_sayisi": len(cfg.get("cariler", {}))})
        if p == "/api/kanit":
            kayitlar = []
            yol = KANIT_LOG()
            if os.path.exists(yol):
                with open(yol, encoding="utf-8") as f:
                    for ln in f:
                        ln = ln.strip()
                        if ln:
                            kayitlar.append(json.loads(ln))
            return 200, _JSON, _jb({"dogrulama": kanit_dogrula(), "kayitlar": kayitlar})
        if p.startswith("/uploads/"):
            dad = os.path.basename(p[len("/uploads/"):])
            yol = os.path.join(_uploads(), dad)
            if os.path.exists(yol):
                ctype = mimetypes.guess_type(yol)[0] or "application/octet-stream"
                with open(yol, "rb") as f:
                    return 200, ctype, f.read()
        return 404, _JSON, _jb({"hata": "bulunamadi"})
    except Exception as e:
        return 500, _JSON, _jb({"hata": str(e)})


def dispatch_post(subpath, body, ip="", ua="", host="", public=None):
    _CTX.host = host
    if public:
        set_aktif(public.get("musteri") or None)
        # Token hangi cari icin imzalandiysa yanit O cariye yazilir.
        body = dict(body)
        body["cari"] = public.get("cari", "")
        body["musteri"] = public.get("musteri", "")
    else:
        set_aktif(body.get("musteri"))
    p = subpath or "/"
    try:
        if public and p not in PORTAL_POST_ACIK:
            return 404, _JSON, _jb({"hata": "bulunamadi"})
        if p == "/api/yanit":
            return 200, _JSON, _jb(kaydet_yanit(body, ip, ua))
        if p == "/api/otp_iste":
            return 200, _JSON, _jb(otp_iste(body.get("cari")))
        if public:
            return 404, _JSON, _jb({"hata": "bulunamadi"})
        if p == "/api/manuel_esle":
            return 200, _JSON, _jb(manuel_esle(body, ip, ua))
        if p == "/api/manuel_geri":
            return 200, _JSON, _jb(manuel_geri(body, ip, ua))
        if p == "/api/gonder":
            return 200, _JSON, _jb(gonder_tur(body.get("secili")))
        if p == "/api/mailayar":
            return 200, _JSON, _jb(api_mailayar_yaz(body, ip, ua))
        if p == "/api/mailtest":
            alici = body.get("alici") or _json_oku(CARI_MAIL(), {}).get("varsayilan_karsi", "")
            html = _mail_govde("TEST", "Test Alicisi", 12345.67)
            ok, hata, mid = mail_gonder(alici, f"Test · Mutabakat · {aktif_unvan()}", html)
            return 200, _JSON, _jb({"ok": ok, "alici": alici, "hata": hata,
                                    "message_id": mid, "mail_aktif": mail_aktif_mi()})
        return 404, _JSON, _jb({"hata": "bulunamadi"})
    except Exception as e:
        return 500, _JSON, _jb({"hata": str(e)})
