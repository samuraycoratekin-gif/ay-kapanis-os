# -*- coding: utf-8 -*-
"""
Akilli Mutabakat Motoru (prototip, cekirdek eslestirme).

Girdi : bizim_ekstreler.xlsx  +  karsi_ekstreler.xlsx
Cikti : konsol triyaj ozeti + Mutabakat_AI/mutabakat_raporu.xlsx

Tasarim ilkesi: her esleme/uyusmazlik GEREKCESIYLE raporlanir; motor asla
"otomatik kayit" atmaz, yalnizca insanin onayina sunulacak oneriyi uretir.

Eslestirme zinciri (cari bazinda):
  1) EXACT        : evrak no (normalize) + tutar birebir
  2) FUZZY        : evrak no cekirdek sayisi ayni / format-tarih farki, tutar ayni
  3) TUTAR FARKI  : ayni belge eslesir ama tutar farkli (fark raporlanir)
  4) SUBSET-SUM   : karsidaki tek odeme, bizdeki cok faturaya dagilmis (kombinasyon)
  5) EKSIK BELGE  : yalnizca bir tarafta olan kayitlar

Bagimlilik: yalnizca openpyxl + stdlib (difflib). numpy/pandas YOK.
"""
import re
from datetime import datetime, date
from difflib import SequenceMatcher
from itertools import combinations
from openpyxl import Workbook, load_workbook

TOL = 0.01          # tutar toleransi (TL)
GUN_TOL = 5         # tarih kaymasi toleransi (gun)
FUZZY_ORAN = 0.85   # difflib benzerlik esigi
SUBSET_MAX = 4      # parcali odemede kombinasyon ust siniri


# --------------------------------------------------------------------------- #
# Yardimcilar
# --------------------------------------------------------------------------- #
def norm_belge(s):
    """Evrak no: buyuk harf, alfanumerik disi temizle."""
    return re.sub(r"[^A-Z0-9]", "", str(s).upper())


def sayi_cekirdek(s):
    """Evrak nodaki son rakam dizisini int olarak dondurur (basindaki sifirlari atar).
    'FT-2026-00318' -> 318 ; '318' -> 318 ; eslesme icin format farkini yutar."""
    rakamlar = re.findall(r"\d+", str(s))
    return int(rakamlar[-1]) if rakamlar else None


def gun_farki(a, b):
    if isinstance(a, datetime):
        a = a.date()
    if isinstance(b, datetime):
        b = b.date()
    if isinstance(a, date) and isinstance(b, date):
        return abs((a - b).days)
    return None


def yakin(x, y, tol=TOL):
    return abs(float(x) - float(y)) <= tol


def tarih_str(d):
    if isinstance(d, (datetime, date)):
        return d.strftime("%d.%m.%Y")
    return str(d)


# --------------------------------------------------------------------------- #
# Veri okuma
# --------------------------------------------------------------------------- #
def oku(dosya):
    wb = load_workbook(dosya, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    basliklar = [str(h) for h in rows[0]]
    kayitlar = []
    for i, r in enumerate(rows[1:]):
        d = dict(zip(basliklar, r))
        d["_id"] = i
        d["_eslesti"] = False
        kayitlar.append(d)
    wb.close()
    return kayitlar


def grupla(kayitlar):
    g = {}
    for k in kayitlar:
        g.setdefault(k["cari_kodu"], []).append(k)
    return g


# --------------------------------------------------------------------------- #
# Eslestirme motoru (tek cari)
# --------------------------------------------------------------------------- #
def cari_esle(bizim, karsi):
    """Tek cari icin eslestirme. Bulgu listesi dondurur."""
    bulgular = []

    def isaretle(b, k):
        b["_eslesti"] = True
        k["_eslesti"] = True

    # 1) EXACT  -------------------------------------------------------------
    for b in bizim:
        if b["_eslesti"]:
            continue
        for k in karsi:
            if k["_eslesti"] or k["tip"] != b["tip"]:
                continue
            if norm_belge(b["belge_no"]) == norm_belge(k["belge_no"]) and \
               yakin(b["tutar"], k["tutar"]):
                isaretle(b, k)
                bulgular.append(("EXACT", b, k,
                                 "Evrak no + tutar birebir aynı"))
                break

    # 2) FUZZY + 3) TUTAR FARKI  -------------------------------------------
    for b in bizim:
        if b["_eslesti"]:
            continue
        en_iyi, en_iyi_skor = None, 0.0
        for k in karsi:
            if k["_eslesti"] or k["tip"] != b["tip"]:
                continue
            cekirdek_ayni = (sayi_cekirdek(b["belge_no"]) is not None and
                             sayi_cekirdek(b["belge_no"]) == sayi_cekirdek(k["belge_no"]))
            benzerlik = SequenceMatcher(
                None, norm_belge(b["belge_no"]), norm_belge(k["belge_no"])).ratio()
            aday = cekirdek_ayni or benzerlik >= FUZZY_ORAN
            if not aday:
                continue
            skor = 1.0 if cekirdek_ayni else benzerlik
            if skor > en_iyi_skor:
                en_iyi, en_iyi_skor = k, skor
        if en_iyi is None:
            continue
        k = en_iyi
        if yakin(b["tutar"], k["tutar"]):
            isaretle(b, k)
            nedenler = []
            if norm_belge(b["belge_no"]) != norm_belge(k["belge_no"]):
                nedenler.append(f"evrak no format farkı ({b['belge_no']} ↔ {k['belge_no']})")
            gf = gun_farki(b["tarih"], k["tarih"])
            if gf and gf > 0:
                nedenler.append(f"{gf} gün tarih kayması")
            bulgular.append(("FUZZY", b, k,
                             "Toleranslı eşleşme: " + (", ".join(nedenler) or "küçük format farkı")))
        else:
            isaretle(b, k)
            fark = round(float(b["tutar"]) - float(k["tutar"]), 2)
            bulgular.append(("TUTAR_FARKI", b, k,
                             f"Aynı belge, tutar farklı: bizde {b['tutar']:.2f} / "
                             f"karşıda {k['tutar']:.2f} → fark {fark:+.2f} TL"))

    # 4) SUBSET-SUM  --------------------------------------------------------
    # Bir taraftaki tek kalemi, diger taraftaki birden cok kalemin toplami kapatiyor mu?
    def subset_dene(tekler, parcalar, tek_taraf):
        for t in tekler:
            if t["_eslesti"]:
                continue
            havuz = [p for p in parcalar if not p["_eslesti"] and p["tip"] == t["tip"]]
            bulundu = None
            for n in range(2, SUBSET_MAX + 1):
                for combo in combinations(havuz, n):
                    if yakin(sum(float(p["tutar"]) for p in combo), t["tutar"]):
                        bulundu = combo
                        break
                if bulundu:
                    break
            if bulundu:
                t["_eslesti"] = True
                kalemler = []
                for p in bulundu:
                    p["_eslesti"] = True
                    kalemler.append(f"{p['belge_no']}={float(p['tutar']):.2f}")
                if tek_taraf == "karsi":
                    aciklama = (f"Karşı tarafın tek ödemesi ({t['belge_no']}={float(t['tutar']):.2f}) "
                                f"bizdeki {len(bulundu)} kalemin toplamı: " + " + ".join(kalemler))
                else:
                    aciklama = (f"Bizdeki tek kalem ({t['belge_no']}={float(t['tutar']):.2f}) "
                                f"karşıdaki {len(bulundu)} kalemin toplamı: " + " + ".join(kalemler))
                bulgular.append(("SUBSET_SUM", t, bulundu, aciklama))

    subset_dene([k for k in karsi if not k["_eslesti"]],
                [b for b in bizim if not b["_eslesti"]], "karsi")
    subset_dene([b for b in bizim if not b["_eslesti"]],
                [k for k in karsi if not k["_eslesti"]], "bizim")

    # 5) EKSIK BELGE  -------------------------------------------------------
    for b in bizim:
        if not b["_eslesti"]:
            bulgular.append(("EKSIK_KARSIDA", b, None,
                             f"Bizde var, karşıda yok: {b['belge_no']} "
                             f"({b['belge_tipi']}) {float(b['tutar']):.2f} TL"))
    for k in karsi:
        if not k["_eslesti"]:
            bulgular.append(("EKSIK_BIZDE", None, k,
                             f"Karşıda var, bizde yok: {k['belge_no']} "
                             f"({k['belge_tipi']}) {float(k['tutar']):.2f} TL"))
    return bulgular


# --------------------------------------------------------------------------- #
# Triyaj
# --------------------------------------------------------------------------- #
SORUNSUZ = {"EXACT", "FUZZY", "SUBSET_SUM"}

def cari_durum(bulgular):
    tipler = {b[0] for b in bulgular}
    if tipler & {"EKSIK_KARSIDA", "EKSIK_BIZDE"}:
        return "EKSIK BELGE"
    if "TUTAR_FARKI" in tipler:
        return "TUTAR FARKLI"
    return "MUTABIK"


def fark_tutari(bulgular):
    """Cari icin net acik (eslesemeyen + tutar farklari), kaba gosterge."""
    toplam = 0.0
    for tip, b, k, _ in bulgular:
        if tip == "TUTAR_FARKI":
            toplam += abs(float(b["tutar"]) - float(k["tutar"]))
        elif tip == "EKSIK_KARSIDA":
            toplam += abs(float(b["tutar"]))
        elif tip == "EKSIK_BIZDE":
            toplam += abs(float(k["tutar"]))
    return round(toplam, 2)


# --------------------------------------------------------------------------- #
# Rapor
# --------------------------------------------------------------------------- #
def rapor_yaz(sonuclar, dosya):
    wb = Workbook()
    ws = wb.active
    ws.title = "Bulgular"
    ws.append(["cari_kodu", "cari_adi", "durum", "bulgu_tipi", "gerekce",
               "bizim_belge", "bizim_tutar", "karsi_belge", "karsi_tutar"])
    for cari_kodu, cari_adi, durum, bulgular in sonuclar:
        for tip, b, k, aciklama in bulgular:
            ws.append([
                cari_kodu, cari_adi, durum, tip, aciklama,
                (b["belge_no"] if b and not isinstance(b, tuple) else ""),
                (float(b["tutar"]) if b and not isinstance(b, tuple) else ""),
                (k["belge_no"] if k and not isinstance(k, (tuple, list)) else
                 (",".join(p["belge_no"] for p in k) if isinstance(k, (tuple, list)) else "")),
                (float(k["tutar"]) if k and not isinstance(k, (tuple, list)) else ""),
            ])
    wb.save(dosya)


# --------------------------------------------------------------------------- #
# Ana akis
# --------------------------------------------------------------------------- #
def main():
    bizim = grupla(oku("Mutabakat_AI/bizim_ekstreler.xlsx"))
    karsi = grupla(oku("Mutabakat_AI/karsi_ekstreler.xlsx"))
    tum_cariler = sorted(set(bizim) | set(karsi))

    sonuclar = []
    sayac = {"MUTABIK": 0, "TUTAR FARKLI": 0, "EKSIK BELGE": 0}
    for ck in tum_cariler:
        bulgular = cari_esle(bizim.get(ck, []), karsi.get(ck, []))
        adi = (bizim.get(ck) or karsi.get(ck))[0]["cari_adi"]
        durum = cari_durum(bulgular)
        sayac[durum] += 1
        sonuclar.append((ck, adi, durum, bulgular))

    toplam = len(tum_cariler)
    mutabik = sayac["MUTABIK"]
    print("=" * 70)
    print("  AKILLI MUTABAKAT - PORTFOY TRIYAJ OZETI")
    print("=" * 70)
    print(f"  Toplam cari            : {toplam}")
    print(f"  Otomatik mutabik       : {mutabik}  (%{100*mutabik/toplam:.0f})")
    print(f"  Tutar farkli (incele)  : {sayac['TUTAR FARKLI']}")
    print(f"  Eksik belge (incele)   : {sayac['EKSIK BELGE']}")
    print(f"  -> Insan gozu gereken  : {toplam - mutabik} cari")
    print("=" * 70)

    print("\n  SORUNLU CARILER (oncelik sirasi):\n")
    for ck, adi, durum, bulgular in sonuclar:
        if durum == "MUTABIK":
            continue
        print(f"  [{durum}] {ck} - {adi}   | tahmini acik: {fark_tutari(bulgular):.2f} TL")
        for tip, b, k, aciklama in bulgular:
            if tip in SORUNSUZ:
                continue
            print(f"       - {aciklama}")
        print()

    print("  MUTABIK CARILER (otomatik kapandi):")
    for ck, adi, durum, bulgular in sonuclar:
        if durum == "MUTABIK":
            n = len(bulgular)
            print(f"       OK {ck} - {adi}  ({n} kalem eslesti)")

    rapor_yaz(sonuclar, "Mutabakat_AI/mutabakat_raporu.xlsx")
    print("\n  Detayli rapor: Mutabakat_AI/mutabakat_raporu.xlsx")


if __name__ == "__main__":
    main()
