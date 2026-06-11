# -*- coding: utf-8 -*-
"""
Mizan okuyucu.

Iki bicimi destekler:
  1) Kumule aylik mizan (ERP GL raporu): Ust Hesap / Hesap / aylik bakiye sutunlari
     (ACILIS, OCAK, SUBAT, ... + Toplam). Aylik sutunlar NET HAREKET, Toplam = kumule
     guncel bakiye. Ornek: "MP GL Kumule Mizan".
  2) Standart mizan: Hesap Kodu / Hesap Adi / Borc / Alacak / Borc Bakiye / Alacak Bakiye.

Cikti: ana hesap (3 haneli) bazinda ozet. Firma/sube adi anonim (Referans okunmaz).

.xls -> xlrd, .xlsx -> openpyxl. (pandas gerekmez.)

Isaret kurali: pozitif = BORC bakiye, negatif = ALACAK bakiye.
"""
import os
from core import metin

# Acilis/devir sutun adlari (hareket degil, acilis bakiyesi)
ACILIS_ADLARI = {"ACILIS", "ACILISBAKIYE", "DEVIR", "DEVIRBAKIYE", "ACILISBAKIYESI"}
# Ay sutunlari (takvim sirasi)
AY_SIRA = ["OCAK", "SUBAT", "MART", "NISAN", "MAYIS", "HAZIRAN",
           "TEMMUZ", "AGUSTOS", "EYLUL", "EKIM", "KASIM", "ARALIK"]
TOPLAM_ADLARI = {"TOPLAM", "GENELTOPLAM", "BAKIYE"}


# Tek dogru kaynak core/metin.py'de; geriye donuk uyum icin alias birakildi.
_nrm = metin.nrm
_sayi = metin.sayi


# --------------------------------------------------------------------------- #
# Ham satir okuma (motor-bagimsiz: liste-of-liste doner)
# --------------------------------------------------------------------------- #
def _ham_satirlar(yol):
    ext = os.path.splitext(yol)[1].lower()
    if ext in (".xlsx", ".xlsm"):
        from openpyxl import load_workbook
        wb = load_workbook(yol, read_only=True, data_only=True)
        try:
            ws = _sayfa_sec_openpyxl(wb)
            return [list(r) for r in ws.iter_rows(values_only=True)]
        finally:
            wb.close()   # read_only kip dosya tutamacini acik birakir (Win kilidi)
    elif ext == ".xls":
        import xlrd
        wb = xlrd.open_workbook(yol)
        ws = _sayfa_sec_xlrd(wb)
        return [[ws.cell_value(r, c) for c in range(ws.ncols)] for r in range(ws.nrows)]
    raise ValueError(f"Desteklenmeyen dosya turu: {ext}")


def _sayfa_sec_openpyxl(wb):
    for ad in wb.sheetnames:
        if "MIZAN" in _nrm(ad):
            return wb[ad]
    return wb[wb.sheetnames[-1]]


def _sayfa_sec_xlrd(wb):
    for ad in wb.sheet_names():
        if "MIZAN" in _nrm(ad):
            return wb.sheet_by_name(ad)
    return wb.sheet_by_index(wb.nsheets - 1)


# --------------------------------------------------------------------------- #
# Baslik satiri + sutun haritasi
# --------------------------------------------------------------------------- #
def _baslik_bul(satirlar):
    """'Hesap' iceren ve veri basligi olan satiri bulur. (satir_index, normalize_hucreler)"""
    for i, r in enumerate(satirlar[:40]):
        nrm = [_nrm(c) for c in r]
        has_hesap = any(c in ("HESAP", "HESAPKODU", "HESAPNO") for c in nrm)
        has_ad = any("HESAP" in c and ("ACIKLAMA" in c or "ADI" in c) for c in nrm)
        if has_hesap and has_ad:
            return i, nrm
    return None, None


def _sutun_haritasi(nrm):
    """Normalize basliklardan sutun indekslerini cikarir."""
    h = {}

    def bul(pred):
        for j, c in enumerate(nrm):
            if pred(c):
                return j
        return None

    h["ana"] = bul(lambda c: c in ("USTHESAP",) or c == "ANAHESAP")
    h["ana_ad"] = bul(lambda c: "USTHESAP" in c and "ACIKLAMA" in c)
    h["kod"] = bul(lambda c: c in ("HESAP", "HESAPKODU", "HESAPNO"))
    h["ad"] = bul(lambda c: "HESAP" in c and ("ACIKLAMA" in c or "ADI" in c) and "UST" not in c)
    # standart mizan sutunlari
    h["borc"] = bul(lambda c: c == "BORC")
    h["alacak"] = bul(lambda c: c == "ALACAK")
    h["borc_bak"] = bul(lambda c: c in ("BORCBAKIYE", "BORCBAKIYESI"))
    h["alacak_bak"] = bul(lambda c: c in ("ALACAKBAKIYE", "ALACAKBAKIYESI"))
    # donem (aylik) sutunlari
    acilis_idx, ay_idx, toplam_idx = None, [], None
    for j, c in enumerate(nrm):
        if c in ACILIS_ADLARI:
            acilis_idx = j
        elif c in AY_SIRA:
            ay_idx.append((AY_SIRA.index(c), j, c))
        elif c in TOPLAM_ADLARI and toplam_idx is None:
            toplam_idx = j
    ay_idx.sort()
    h["acilis"] = acilis_idx
    h["aylar"] = [(c, j) for (_, j, c) in ay_idx]   # [(ay_adi, sutun_idx), ...]
    h["toplam"] = toplam_idx
    return h


# --------------------------------------------------------------------------- #
# Ana okuma
# --------------------------------------------------------------------------- #
def oku(yol):
    satirlar = _ham_satirlar(yol)
    hi, nrm = _baslik_bul(satirlar)
    if hi is None:
        raise ValueError("Mizan basligi taninamadi (Hesap/Hesap Aciklama sutunlari yok).")
    h = _sutun_haritasi(nrm)
    if h["kod"] is None:
        raise ValueError("Hesap kodu sutunu bulunamadi.")

    aylar = [a for (a, _) in h["aylar"]]
    kumule_format = bool(h["aylar"]) or h["toplam"] is not None

    # --- 1. gecis: satirlari ayristir (henuz toplama yok) -------------------
    parsed = []
    for r in satirlar[hi + 1:]:
        kod = r[h["kod"]] if h["kod"] < len(r) else None
        if kod in (None, ""):
            continue
        kod = str(kod).strip()
        if not kod or not kod[0].isdigit():
            continue
        ad = str(r[h["ad"]]).strip() if (h["ad"] is not None and h["ad"] < len(r) and r[h["ad"]]) else ""

        # ana hesap (3 haneli)
        if h["ana"] is not None and h["ana"] < len(r) and r[h["ana"]] not in (None, ""):
            ana = str(r[h["ana"]]).strip()
        else:
            ana = kod[:3]
        ana_ad = ""
        if h["ana_ad"] is not None and h["ana_ad"] < len(r) and r[h["ana_ad"]]:
            ana_ad = str(r[h["ana_ad"]]).strip()

        # bakiye / aylik
        if kumule_format:
            acilis = _sayi(r[h["acilis"]]) if (h["acilis"] is not None and h["acilis"] < len(r)) else None
            aylik = {}
            for ay, j in h["aylar"]:
                aylik[ay] = _sayi(r[j]) if j < len(r) else None
            if h["toplam"] is not None and h["toplam"] < len(r):
                toplam = _sayi(r[h["toplam"]])
            else:
                toplam = (acilis or 0) + sum(v or 0 for v in aylik.values())
        else:
            bb = _sayi(r[h["borc_bak"]]) if h["borc_bak"] is not None and h["borc_bak"] < len(r) else None
            ab = _sayi(r[h["alacak_bak"]]) if h["alacak_bak"] is not None and h["alacak_bak"] < len(r) else None
            if bb is None and ab is None:
                # borc/alacak hareketten net
                bb = _sayi(r[h["borc"]]) if h["borc"] is not None and h["borc"] < len(r) else 0
                ab = _sayi(r[h["alacak"]]) if h["alacak"] is not None and h["alacak"] < len(r) else 0
            toplam = (bb or 0) - (ab or 0)
            acilis, aylik = None, {}

        parsed.append({"kod": kod, "ad": ad, "ana": ana, "ana_ad": ana_ad,
                       "toplam": 0.0 if toplam is None else toplam,
                       "acilis": acilis, "aylik": aylik})

    # Hem 3 haneli GRUP satiri (kod == ana, or. "770") hem detaylari ("770.01")
    # olan mizanlarda grup satiri TOPLAMA katilmaz (cift sayim onlenir);
    # yalnizca grup ADI icin kullanilir.
    detayli_analar = {p["ana"] for p in parsed if p["kod"] != p["ana"]}

    # --- 2. gecis: ana hesap bazinda topla ---------------------------------
    hesaplar = {}   # ana_kod -> ozet
    satir_sayisi = 0
    for p in parsed:
        ana, kod, ad, ana_ad = p["ana"], p["kod"], p["ad"], p["ana_ad"]
        o = hesaplar.setdefault(ana, {"ana": ana, "ad": ana_ad,
                                      "toplam": 0.0, "acilis": 0.0,
                                      "aylik": {a: 0.0 for a in aylar},
                                      "detay_sayisi": 0})
        # Grup adi: once ust-hesap aciklamasi; yoksa ana hesabin kendi satiri
        # (kod == ana, or. "770"); o da yoksa ilk detay satirinin adi.
        if ana_ad:
            o["ad"] = ana_ad
        elif kod == ana and ad:
            o["ad"] = ad
        elif not o["ad"] and ad:
            o["ad"] = ad
        if kod == ana and ana in detayli_analar:
            continue                     # ozet satiri: tutarlari detaylar tasiyor
        o["toplam"] += p["toplam"]
        o["acilis"] += (p["acilis"] or 0)
        for a in aylar:
            o["aylik"][a] += (p["aylik"].get(a) or 0)
        o["detay_sayisi"] += 1
        satir_sayisi += 1

    return {
        "kaynak": os.path.basename(yol),
        "format": "kumule" if kumule_format else "standart",
        "aylar": aylar,
        "guncel_ay": aylar[-1] if aylar else None,
        "hesap_sayisi": len(hesaplar),
        "satir_sayisi": satir_sayisi,
        "hesaplar": hesaplar,
    }
