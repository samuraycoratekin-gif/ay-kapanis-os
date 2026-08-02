# -*- coding: utf-8 -*-
"""
Optimed Saglik Grubu -- Ornek Mutabakat Verisi Ureteci
Haziran 2026  |  Logo Tiger <-> SGK/Medula <-> GIB

3 dosya uretir:
  optimed_logo.xlsx  -- Bizim  (Logo Tiger ERP cari ekstreleri)
  optimed_sgk.xlsx   -- Karsi  (SGK + Sigorta + Kurumsal kayitlari)
  optimed_gib.xlsx   -- GIB   (e-Fatura hakem kayitlari)

10 senaryo:
  #01  EXACT          SGK/Cerkezkoy -- diyaliz, temiz esleme
  #02  TUTAR_FARKI    SGK/Cerkezkoy -- S1: cerrahi SGK kesintisi (12.490 TL)
  #03  FUZZY          SGK/Cerkezkoy -- poliklinik, evrak no format farki
  #04  SUBSET_SUM     SGK/Cerkezkoy -- 3 avans odemesi = 1 toplu SGK dekontu
  #05  EKSIK_KARSIDA  SGK/Corlu     -- S2: askida gelir (provizyon alinamadi)
  #06  EXACT          SGK/Kapakli   -- acil servis, temiz
  #07  FUZZY          Sigorta/ACI   -- Acibademin referans no farki
  #08  EKSIK_BIZDE    Sigorta/ACI   -- S8: split hatasi (sigortada var, bizde yok)
  #09  EXACT          Kurumsal/OSB  -- isyeri hekimligi, temiz
  #10  EKSIK_KARSIDA  Kurumsal/OSB  -- S5: kurumsal kacak (karsida kayit yok)
"""
import os
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# --------------------------------------------------------------------------- #
# Yollar
# --------------------------------------------------------------------------- #
BURADA = os.path.dirname(os.path.abspath(__file__))
HEDEF  = os.path.join(BURADA, "moduller", "mutabakat", "ornek_veri")

LOGO_SUTUNLAR = ["cari_kodu", "cari_adi", "tarih", "belge_no",
                 "belge_tipi", "aciklama", "tutar", "tip"]
GIB_SUTUNLAR  = ["cari_kodu", "cari_adi", "tarih", "belge_no",
                 "belge_tipi", "tutar"]

# --------------------------------------------------------------------------- #
# Cari tanimlari
# --------------------------------------------------------------------------- #
CARI = {
    "120.01.001": "SGK - Cerkezkoy Merkez Hastanesi",
    "120.01.002": "SGK - Corlu International Hastanesi",
    "120.01.003": "SGK - Kapakli Hastanesi",
    "120.02.001": "Acibademin Tamamlayici Sigorta A.S.",
    "120.03.001": "Kurumsal Alacaklar - Cerkezkoy OSB",
}

# --------------------------------------------------------------------------- #
# Tarih sabitleri
# --------------------------------------------------------------------------- #
HAZ01 = date(2026, 6,  1)
HAZ05 = date(2026, 6,  5)
HAZ10 = date(2026, 6, 10)
HAZ15 = date(2026, 6, 15)
HAZ20 = date(2026, 6, 20)
HAZ22 = date(2026, 6, 22)
HAZ25 = date(2026, 6, 25)
HAZ27 = date(2026, 6, 27)
HAZ28 = date(2026, 6, 28)
HAZ30 = date(2026, 6, 30)
TEM01 = date(2026, 7,  1)


# --------------------------------------------------------------------------- #
# Yardimcilar
# --------------------------------------------------------------------------- #
def satir(cari_kodu, tarih, belge_no, belge_tipi, aciklama, tutar,
          tip="FATURA"):
    return {
        "cari_kodu":  cari_kodu,
        "cari_adi":   CARI[cari_kodu],
        "tarih":      tarih,
        "belge_no":   belge_no,
        "belge_tipi": belge_tipi,
        "aciklama":   aciklama,
        "tutar":      float(tutar),
        "tip":        tip,
    }


def gib_satir(cari_kodu, tarih, belge_no, belge_tipi, tutar):
    return {
        "cari_kodu":  cari_kodu,
        "cari_adi":   CARI[cari_kodu],
        "tarih":      tarih,
        "belge_no":   belge_no,
        "belge_tipi": belge_tipi,
        "tutar":      float(tutar),
    }


# --------------------------------------------------------------------------- #
# LOGO SATIRLARI  (bizim -- Logo Tiger ERP)
# --------------------------------------------------------------------------- #
LOGO_SATIRLAR = [

    # === SGK / Cerkezkoy (120.01.001) ===

    # 01 EXACT -- diyaliz paketi, tam esleme bekleniyor
    satir("120.01.001", HAZ30, "FT-CRK-2026-0612-01", "e-Fatura",
          "Haz.2026 diyaliz paketi", 84_250.00),

    # 02 TUTAR_FARKI -- cerrahi; SGK 12.490 TL keser (SUT kapsami disi)
    satir("120.01.001", HAZ28, "FT-CRK-2026-0612-02", "e-Fatura",
          "Haz.2026 cerrahi islemler (tam tutar)", 127_840.00),

    # 03 FUZZY -- poliklinik; SGK evrak no kisaltmis ("0612-03")
    satir("120.01.001", HAZ25, "FT-CRK-2026-0612-03", "e-Fatura",
          "Haz.2026 poliklinik ve muayene", 18_620.00),

    # 04 SUBSET_SUM -- 3 ayri avans dekontu; SGK tek toplu mahsupla kapatir
    satir("120.01.001", HAZ05, "ODM-CRK-2026-0601", "Dekont",
          "SGK Haz.2026 avans 1/3", 35_000.00, tip="ODEME"),
    satir("120.01.001", HAZ15, "ODM-CRK-2026-0602", "Dekont",
          "SGK Haz.2026 avans 2/3", 45_000.00, tip="ODEME"),
    satir("120.01.001", HAZ25, "ODM-CRK-2026-0603", "Dekont",
          "SGK Haz.2026 avans 3/3", 22_500.00, tip="ODEME"),

    # === SGK / Corlu (120.01.002) ===

    # 05 EKSIK_KARSIDA -- S2 askida gelir; Medula provizyon almamis
    satir("120.01.002", HAZ28, "FT-COR-2026-0611-01", "e-Fatura",
          "Haz.2026 yatakli servis - provizyon bekliyor", 54_320.00),

    # === SGK / Kapakli (120.01.003) ===

    # 06 EXACT -- acil servis, temiz
    satir("120.01.003", HAZ20, "FT-KPK-2026-0611-01", "e-Fatura",
          "Haz.2026 acil servis hizmetleri", 8_750.00),

    # === Tamamlayici Sigorta (120.02.001) ===

    # 07 FUZZY -- Acibademin kendi ref formati: "45/2026/ACI"
    satir("120.02.001", HAZ20, "FT-CRK-SIG-2026-0045", "e-Arsiv",
          "Tamamlayici sigorta - ameliyat farki", 4_320.00),

    # 08 EKSIK_BIZDE -- S8 split hatasi; sigorta kaydi var, bizde kayit yok
    #    Logo satirina girmez; yalnizca SGK dosyasinda yer alir.

    # === Kurumsal / OSB (120.03.001) ===

    # 09 EXACT -- isyeri hekimligi Mayis donemi
    satir("120.03.001", HAZ15, "FT-KUR-2026-0611-01", "e-Arsiv",
          "Cerkezkoy OSB isyeri hekimligi - Mayis", 28_500.00),

    # 10 EKSIK_KARSIDA -- S5 kurumsal kacak; karsida kayit yok
    satir("120.03.001", HAZ20, "FT-KUR-2026-0611-02", "e-Arsiv",
          "Cerkezkoy OSB isyeri hekimligi - Haziran", 14_250.00),
]

# --------------------------------------------------------------------------- #
# SGK SATIRLARI  (karsi taraf -- SGK/Medula + Sigorta + Kurumsal)
# --------------------------------------------------------------------------- #
SGK_SATIRLAR = [

    # 01 EXACT
    satir("120.01.001", HAZ30, "FT-CRK-2026-0612-01", "e-Fatura",
          "Diyaliz - tam kabul", 84_250.00),

    # 02 TUTAR_FARKI -- SGK 12.490 TL keserek 115.350 TL odedi
    satir("120.01.001", HAZ28, "FT-CRK-2026-0612-02", "e-Fatura",
          "Cerrahi - SUT kapsami disi 12.490 TL kesildi", 115_350.00),

    # 03 FUZZY -- SGK kisaltilmis evrak no kullanmis
    satir("120.01.001", HAZ27, "0612-03", "e-Fatura",
          "Poliklinik - kabul", 18_620.00),

    # 04 SUBSET_SUM -- tek toplu SGK avans dekontu (35+45+22.5 = 102.5K)
    satir("120.01.001", TEM01, "DEK-2026-CRK-001", "Dekont",
          "Haz.2026 SGK avans mahsubu (toplam)", 102_500.00, tip="ODEME"),

    # 05 EKSIK_KARSIDA -- yok (SGK'da kayit bulunmuyor)

    # 06 EXACT
    satir("120.01.003", HAZ20, "FT-KPK-2026-0611-01", "e-Fatura",
          "Acil servis - tam kabul", 8_750.00),

    # 07 FUZZY -- Acibademin kendi ref numarasi (cekirdek=45, Logo'daki 0045 ile eslenir)
    satir("120.02.001", HAZ22, "ACI-SIG-2026-0045", "e-Arsiv",
          "Ameliyat farki odemesi", 4_320.00),

    # 08 EKSIK_BIZDE -- sigortada kayit var, Logo'da yok
    satir("120.02.001", HAZ10, "AC-2026-00892", "e-Arsiv",
          "Hasta payi - bizdeki kayit eksik (split hatasi)", 2_180.00),

    # 09 EXACT
    satir("120.03.001", HAZ15, "FT-KUR-2026-0611-01", "e-Arsiv",
          "OSB isyeri hekimligi - Mayis kabul", 28_500.00),

    # 10 EKSIK_KARSIDA -- yok
]

# --------------------------------------------------------------------------- #
# GIB SATIRLARI  (hakem -- e-Fatura + e-Arsiv kayitlari)
# NOT: Dekont/ODEME tipi GIB'de yer almaz; yalnizca fatura tipi satirlar.
# --------------------------------------------------------------------------- #
GIB_SATIRLAR = [
    # SADECE e-Fatura kayitlari: GIB modulu b_ef/k_ef'i "e-Fatura" filtresiyle
    # olusturur; e-Arsiv buraya girerse ASAMA0/1 yanlis-pozitif verir.
    # e-Arsiv + Dekont kayitlari ASAMA 2 (ekstre bilateral) ile kontrol edilir.

    # 01 temiz -- diyaliz
    gib_satir("120.01.001", HAZ30, "FT-CRK-2026-0612-01", "e-Fatura", 84_250.00),

    # 02 GIB = Logo tutari (127.840) -> ASAMA1 KARSI: SGK eksik odedi
    gib_satir("120.01.001", HAZ28, "FT-CRK-2026-0612-02", "e-Fatura", 127_840.00),

    # 03 temiz -- poliklinik
    gib_satir("120.01.001", HAZ25, "FT-CRK-2026-0612-03", "e-Fatura", 18_620.00),

    # 05 GIB kaydi var, SGK islememis -> ASAMA1 KARSI: askida gelir kesin kaniti
    gib_satir("120.01.002", HAZ28, "FT-COR-2026-0611-01", "e-Fatura", 54_320.00),

    # 06 temiz -- acil servis
    gib_satir("120.01.003", HAZ20, "FT-KPK-2026-0611-01", "e-Fatura", 8_750.00),

    # Notlar -- e-Arsiv kalemleri GIB dosyasina girmez:
    #   07 FT-CRK-SIG-2026-0045 (e-Arsiv) -> ASAMA2 ekstre ile cozulur
    #   09 FT-KUR-2026-0611-01  (e-Arsiv) -> ASAMA2 ekstre ile cozulur
    #   10 FT-KUR-2026-0611-02  (e-Arsiv) -> ASAMA2 ekstre ile cozulur
]


# --------------------------------------------------------------------------- #
# Excel yazici
# --------------------------------------------------------------------------- #
BASLIK_RENK = "2E5DA6"   # Optimed koyu mavi
BASLIK_YZ   = "FFFFFF"
CIFT_SIRA   = "EBF2FB"   # acik mavi serit


def yaz_xlsx(yol, satirlar, sutunlar, sayfa_adi):
    wb = Workbook()
    ws = wb.active
    ws.title = sayfa_adi

    ws.append(sutunlar)
    for hucre in ws[1]:
        hucre.font      = Font(bold=True, color=BASLIK_YZ)
        hucre.fill      = PatternFill("solid", fgColor=BASLIK_RENK)
        hucre.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20

    for idx, r in enumerate(satirlar, 2):
        deger = [r.get(s) for s in sutunlar]
        ws.append(deger)
        if idx % 2 == 0:
            for hucre in ws[idx]:
                hucre.fill = PatternFill("solid", fgColor=CIFT_SIRA)

    for col in ws.columns:
        en = max(len(str(c.value or "")) for c in col) + 3
        ws.column_dimensions[col[0].column_letter].width = min(en, 44)

    wb.save(yol)
    print(f"  OK  {os.path.basename(yol):30s}  {len(satirlar)} satir")


# --------------------------------------------------------------------------- #
# Ana akis
# --------------------------------------------------------------------------- #
if __name__ == "__main__":
    os.makedirs(HEDEF, exist_ok=True)
    print(f"\nOptimed ornek veri uretiliyor\n  Hedef: {HEDEF}\n")

    yaz_xlsx(os.path.join(HEDEF, "optimed_logo.xlsx"),
             LOGO_SATIRLAR, LOGO_SUTUNLAR, "Logo Tiger ERP")

    yaz_xlsx(os.path.join(HEDEF, "optimed_sgk.xlsx"),
             SGK_SATIRLAR, LOGO_SUTUNLAR, "SGK-Medula")

    yaz_xlsx(os.path.join(HEDEF, "optimed_gib.xlsx"),
             GIB_SATIRLAR, GIB_SUTUNLAR, "GIB e-Fatura")

    print(f"\n3 dosya hazir. 10 senaryo | 5 cari | 3 hastane.\n")
